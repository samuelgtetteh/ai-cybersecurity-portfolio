"""
Builds an Excel workbook with a real formatted Table object (banded rows,
header styling, filter dropdowns — not just a plain CSV opened in Excel),
plus a separate summary sheet for the executive summary text.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import report_export

PRIORITY_FILLS = {
    "Critical": PatternFill(start_color="F8CBCB", end_color="F8CBCB", fill_type="solid"),
    "High": PatternFill(start_color="FCE4C4", end_color="FCE4C4", fill_type="solid"),
    "Medium": PatternFill(start_color="FCF3C4", end_color="FCF3C4", fill_type="solid"),
    "Low": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
}

COLUMN_WIDTHS = {
    "source": 14, "host_ip": 14, "category": 16, "control_id": 12,
    "control_text": 50, "priority": 10, "adjusted_score": 10,
    "reasons": 50, "draft_language": 60, "status": 14, "notes": 30,
}


def build_report(path, final_report, baseline_recommendations, executive_summary, drafts=None, business_name=None):
    rows = report_export.to_rows(final_report, baseline_recommendations, drafts=drafts)
    fieldnames = ["source", "host_ip", "category", "control_id", "control_text",
                  "priority", "adjusted_score", "reasons", "draft_language", "status", "notes"]

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = business_name or "Executive Summary"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = executive_summary
    summary_ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    summary_ws.merge_cells("A2:H20")
    summary_ws.column_dimensions["A"].width = 100
    summary_ws["A1"].alignment = Alignment(wrap_text=False)

    ws = wb.create_sheet("Controls")
    ws.append(fieldnames)
    for row in rows:
        ws.append([row[f] for f in fieldnames])

    for i, field in enumerate(fieldnames, start=1):
        ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(field, 16)

    for row_idx in range(2, len(rows) + 2):
        priority = ws.cell(row=row_idx, column=fieldnames.index("priority") + 1).value
        fill = PRIORITY_FILLS.get(priority)
        for col_idx in range(1, len(fieldnames) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    last_col = get_column_letter(len(fieldnames))
    last_row = len(rows) + 1
    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName="ControlsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    wb.save(path)
    return len(rows)
