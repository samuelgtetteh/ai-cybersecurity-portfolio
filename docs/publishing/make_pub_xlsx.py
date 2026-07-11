"""
Generate the publishing-strategy workbook for the AI-cybersecurity portfolio.

Reproducible: run with the project venv ->
    venv\\Scripts\\python.exe scratchpad\\make_pub_xlsx.py
Output: docs/publishing/publishing_recommendations.xlsx

Sources (verified 2026-07-11 via web search):
  USENIX open access / no APC      https://www.usenix.org/conference/usenixsecurity26/call-for-papers
  USENIX CSET (reproducibility)    https://cset24.isi.edu/
  TMLR (diamond OA)                https://doaj.org/toc/2835-8856
  ACM DTRAP (APC not required)     https://dl.acm.org/journal/dtrap/author-guidelines
  JSys (diamond OA)                https://www.jsys.org/about
  Oxford J. Cybersecurity waiver   https://academic.oup.com/pages/open-research/open-access/charges-licences-and-self-archiving/apc-waiver-policy
  DePaul-Cambridge OA agreement    https://news.library.depaul.press/full-text/2025/10/02/library-announces-open-access-publishing-agreement-with-cambridge-university-press/
  DOAJ / Beall's List (safety)     https://doaj.org/  |  https://beallslist.net/
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- shared styles -------------------------------------------------------
NAVY   = "1F3864"
BLUE   = "2E5496"
LIGHT  = "D9E1F2"
GREEN  = "C6EFCE"
YELLOW = "FFEB9C"
RED    = "FFC7CE"
GREY   = "F2F2F2"

title_font  = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
hdr_font    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell_font   = Font(name="Calibri", size=11, color="000000")
bold_cell   = Font(name="Calibri", size=11, bold=True, color="000000")
note_font   = Font(name="Calibri", size=10, italic=True, color="595959")

title_fill  = PatternFill("solid", fgColor=NAVY)
hdr_fill    = PatternFill("solid", fgColor=BLUE)
light_fill  = PatternFill("solid", fgColor=LIGHT)
grey_fill   = PatternFill("solid", fgColor=GREY)

thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap_top  = Alignment(wrap_text=True, vertical="top")
wrap_ctr  = Alignment(wrap_text=True, vertical="center", horizontal="center")
left_top  = Alignment(wrap_text=True, vertical="top", horizontal="left")

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap_ctr
        cell.border = border

def banner(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 28

def write_table(ws, start_row, headers, rows, widths, fills=None):
    style_header_row_at = start_row
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    ws.row_dimensions[start_row].height = 30
    r = start_row + 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = cell_font
            cell.alignment = left_top
            cell.border = border
            if fills and fills[i]:
                cell.fill = PatternFill("solid", fgColor=fills[i])
            elif i % 2 == 1:
                cell.fill = grey_fill
        ws.row_dimensions[r].height = 46
        r += 1
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return r

wb = Workbook()

# =========================================================================
# Sheet 1: Venue Comparison
# =========================================================================
ws = wb.active
ws.title = "Venue Comparison"
banner(ws, "Publishing Venues — Vetted, Low/No-Cost Options (verified 2026-07-11)", 7)
ws.cell(row=2, column=1,
        value="CS's most reputable venues charge $0 to publish. NOTE: DePaul's CONFIRMED Full Cambridge APC waiver also makes several "
              "indexed Cambridge journals free — see the 'Cambridge OA' sheet (often the best option per paper).")
ws.cell(row=2, column=1).font = note_font
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
ws.row_dimensions[2].height = 30

headers = ["Venue", "Type", "Cost to publish", "Speed", "Legitimacy", "Best-fit paper", "Link"]
rows = [
    ["USENIX CSET", "Workshop, open access",
     "$0 APC (registration only; waivers exist)", "Annual cycle; fast review",
     "USENIX-cooperated, run by USC-ISI", "OT/ICS (top pick)",
     "https://cset24.isi.edu/"],
    ["TMLR (Transactions on Machine Learning Research)", "Diamond OA journal",
     "$0", "Rolling, ~2 months to decision",
     "Highly respected in ML; reviews correctness, not novelty", "Identity (top pick); OT/ICS",
     "https://doaj.org/toc/2835-8856"],
    ["ACM DTRAP (Digital Threats: Research & Practice)", "Gold OA journal",
     "APC requested but NOT required -> effectively $0", "Rolling",
     "ACM, peer-reviewed, industry/threat focus", "All three (strong default)",
     "https://dl.acm.org/journal/dtrap/author-guidelines"],
    ["JSys (Journal of Systems Research)", "Diamond OA journal",
     "$0", "~1.5 months to first decision",
     "Scholarly-led; open peer review; DBLP-indexed", "OT/ICS, Identity",
     "https://www.jsys.org/about"],
    ["arXiv", "Preprint (not peer-reviewed)",
     "$0", "Days",
     "Establishes priority; citable; not a substitute for peer review", "All three (post preprints)",
     "https://arxiv.org/"],
    ["USENIX Security / WOOT", "Conference / workshop, OA",
     "$0 APC", "Fixed deadlines",
     "Top-tier (Security main track very competitive)", "Stretch option",
     "https://www.usenix.org/conference/usenixsecurity26/call-for-papers"],
    ["Oxford Journal of Cybersecurity", "Hybrid journal",
     "APC ~$2k+; waiver = low-income countries only (US NOT eligible)", "Slow",
     "Reputable", "Not recommended on cost",
     "https://academic.oup.com/pages/open-research/open-access/charges-licences-and-self-archiving/apc-waiver-policy"],
]
fills = [GREEN, GREEN, GREEN, GREEN, LIGHT, YELLOW, RED]
end = write_table(ws, 4, headers, rows, widths=[34, 22, 30, 22, 34, 24, 46], fills=fills)
# hyperlink the link column
for i in range(len(rows)):
    r = 5 + i
    cell = ws.cell(row=r, column=7)
    url = rows[i][6]
    cell.hyperlink = url
    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
ws.freeze_panes = "A5"

# =========================================================================
# Sheet 2: Per-Paper Recommendations
# =========================================================================
ws2 = wb.create_sheet("Per-Paper Plan")
banner(ws2, "Per-Paper Recommendations", 5)
headers2 = ["Paper", "Character", "Primary venue", "Backup venue(s)", "Why"]
rows2 = [
    ["OT/ICS — \"Label Leakage in ICS Anomaly Detection\" (HAI testbed)",
     "ML anomaly detection on industrial sensor data / reproducibility",
     "Data-Centric Engineering (Cambridge, $0, indexed)",
     "USENIX CSET; ACM DTRAP; JSys",
     "Cambridge OA journal, FREE for DePaul, and Scopus/WoS-indexed. Already publishes ICS/sensor anomaly detection (e.g. 'Anomaly detection in a fleet of industrial assets', PCA-autoencoder papers) — a dead-on topical fit. Verify DCE's current Scopus status (newer journal, DOAJ-listed)."],
    ["Identity — \"Access Breadth as a Robust Signal for Lateral Movement\" (LANL)",
     "ML feature-attribution / ablation study",
     "TMLR (diamond OA, $0)",
     "ACM DTRAP; Network Science (Cambridge, $0)",
     "TMLR judges technical correctness and clarity, not novelty — ideal for a rigorous ablation result, and free. Network Science is a free Cambridge fallback but a weaker fit (this paper is deliberately non-graph/tabular)."],
    ["RegMap — NIST 800-53 -> HIPAA retrieval (once expanded)",
     "Regulatory NLP / information retrieval",
     "Natural Language Processing (Cambridge, $0, indexed)",
     "Data & Policy (Cambridge, $0); ACM DTRAP",
     "Cambridge NLP journal (formerly Natural Language Engineering), FREE for DePaul, Scopus+WoS+JCR indexed; scope explicitly covers information retrieval. Data & Policy is the compliance/governance-framed alternative (also free, indexed)."],
]
end2 = write_table(ws2, 2, headers2, rows2, widths=[40, 30, 22, 26, 52], fills=[LIGHT, LIGHT, LIGHT])
for r in range(3, 3 + len(rows2)):
    ws2.row_dimensions[r].height = 92
note = ws2.cell(row=end2 + 1, column=1,
    value="Strategy: publishing the three in DISTINCT reputable venues + an arXiv preprint of each is a stronger portfolio signal than clustering them in one place.")
note.font = note_font
ws2.merge_cells(start_row=end2 + 1, start_column=1, end_row=end2 + 1, end_column=5)
ws2.row_dimensions[end2 + 1].height = 30

# =========================================================================
# Sheet 3: Action Items
# =========================================================================
ws3 = wb.create_sheet("Action Items")
banner(ws3, "Action Items & Institutional Levers", 4)
headers3 = ["#", "Action", "Why it matters", "Status"]
rows3 = [
    ["1", "Email DePaul Scholarly Communications Librarian: \"Is DePaul part of ACM Open?\"",
     "If yes, ACM DTRAP and all ACM venues become automatically $0 for you.", "To do"],
    ["2", "Use the CONFIRMED DePaul-Cambridge Full APC waiver (checked 2026-07-11 via CUP Eligibility Checker)",
     "DePaul = 'Full' APC discount: publish Gold OA at $0 in 380+ Cambridge journals. Relevant CS titles: Data-Centric Engineering, Natural Language Processing, Data & Policy, Network Science. See 'Cambridge OA' sheet.", "Confirmed"],
    ["3", "Confirm USENIX CSET next submission deadline",
     "CSET runs on an annual cycle; timing determines when the OT/ICS paper can go out.", "To do"],
    ["4", "Post arXiv preprints of all three papers (OT/ICS blocked on cs.CR endorsement, code PPTQJN)",
     "Establishes priority and citability now; pursue endorser Zhen Huang (DePaul, zhen.huang@depaul.edu).", "In progress"],
    ["5", "Submit one venue at a time (no concurrent journal submissions)",
     "Concurrent submission violates venue policy; arXiv preprints are generally allowed alongside.", "Ongoing"],
]
end3 = write_table(ws3, 2, headers3, rows3, widths=[5, 52, 52, 14], fills=None)
for r in range(3, 3 + len(rows3)):
    ws3.row_dimensions[r].height = 46
    ws3.cell(row=r, column=1).alignment = wrap_ctr

# =========================================================================
# Sheet 4: Safety Checklist
# =========================================================================
ws4 = wb.create_sheet("Safety Checklist")
banner(ws4, "Predatory-Venue Safety Check (run before any journal submission)", 3)
headers4 = ["Step", "How", "Link"]
rows4 = [
    ["Check DOAJ", "Search the journal's ISSN; every DOAJ journal is vetted as non-predatory.", "https://doaj.org/"],
    ["Check Beall's List", "Search the publisher name; if listed, do NOT submit.", "https://beallslist.net/"],
    ["Verify indexing", "Confirm Scopus / DBLP indexing directly (don't trust website logos).", "https://dblp.org/"],
    ["Note for diamond OA", "TMLR and JSys are newer and may not be Scopus-indexed yet, but are DBLP-indexed and well-regarded in CS — which is what matters in our field.", ""],
]
end4 = write_table(ws4, 2, headers4, rows4, widths=[24, 66, 34], fills=[GREEN, GREEN, LIGHT, YELLOW])
for i in range(len(rows4)):
    r = 3 + i
    ws4.row_dimensions[r].height = 44
    url = rows4[i][2]
    if url:
        cell = ws4.cell(row=r, column=3)
        cell.hyperlink = url
        cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

# =========================================================================
# Sheet 5: Cambridge OA (free at DePaul)
# =========================================================================
ws5 = wb.create_sheet("Cambridge OA")
banner(ws5, "DePaul x Cambridge University Press — Full APC Waiver ($0 Gold OA)", 5)
ws5.cell(row=2, column=1,
    value="Confirmed 2026-07-11 via Cambridge Eligibility Checker: DePaul = 'Article Processing Charge Discount: Full'. "
          "Publish Gold OA at NO cost in 380+ Cambridge journals. Eligibility is by CORRESPONDING-AUTHOR affiliation (use DePaul + ORCID). "
          "Relevant CS/ML/data titles below; the rest of the list is humanities/law/medicine.")
ws5.cell(row=2, column=1).font = note_font
ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws5.row_dimensions[2].height = 46
headers5 = ["Cambridge journal", "Fit for your papers", "Indexing", "Notes", "Link"]
rows5 = [
    ["Data-Centric Engineering", "OT/ICS — ideal",
     "DOAJ; verify current Scopus", "Open access; publishes ML anomaly detection on industrial sensor data (fleet anomaly detection, PCA-autoencoder, multivariate time series). Newer journal (2020).",
     "https://www.cambridge.org/core/journals/data-centric-engineering"],
    ["Natural Language Processing (formerly Natural Language Engineering)", "RegMap — ideal",
     "Scopus + Web of Science + JCR", "OA since 1995; scope explicitly includes information retrieval, question answering, text summarisation. Established, reputable.",
     "https://www.cambridge.org/core/journals/natural-language-processing"],
    ["Data & Policy", "RegMap — strong alternative",
     "Scopus", "OA; data science for governance/compliance; privacy/law; publishes replication studies. Compliance-mapping framing fits.",
     "https://www.cambridge.org/core/journals/data-and-policy"],
    ["Network Science", "Identity — possible, weaker fit",
     "Scopus (Q2)", "Cambridge; network/graph-theory oriented. Your identity paper is deliberately non-graph (tabular feature-attribution), so fit is a stretch.",
     "https://www.cambridge.org/core/journals/network-science"],
    ["Cambridge Forum on AI: Law & Governance", "Marginal (RegMap-adjacent)",
     "Newer forum", "AI law/governance forum; not a fit for an empirical ML method paper.",
     "https://www.cambridge.org/core/journals"],
    ["AI EDAM; Mathematical Structures in Computer Science", "Not a fit",
     "Scopus", "AI-for-engineering-design and theoretical CS respectively — out of scope for these applied papers.",
     "https://www.cambridge.org/core/journals"],
]
fills5 = [GREEN, GREEN, GREEN, YELLOW, GREY, GREY]
end5 = write_table(ws5, 3, headers5, rows5, widths=[42, 30, 26, 58, 40], fills=fills5)
for i in range(len(rows5)):
    r = 4 + i
    ws5.row_dimensions[r].height = 62
    url = rows5[i][4]
    cell = ws5.cell(row=r, column=5)
    cell.hyperlink = url
    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
note5 = ws5.cell(row=end5 + 1, column=1,
    value="Source: Cambridge Eligibility Checker (https://www.cambridge.org/core/eligibility-checker), DePaul / United States, 2026-07-11.")
note5.font = note_font
ws5.merge_cells(start_row=end5 + 1, start_column=1, end_row=end5 + 1, end_column=5)

# =========================================================================
# Sheet 6: Processing Time
# =========================================================================
ws6 = wb.create_sheet("Processing Time")
banner(ws6, "Processing / Review Times (checked 2026-07-11)", 5)
ws6.cell(row=2, column=1,
    value="Two things kept separate: time-to-first-decision (when you hear back) vs accepted->online (all are open access, so they post "
          "as soon as accepted - no wait for a print issue). Figures EXCLUDE revision rounds. Realistic journal end-to-end with one "
          "minor-revision round is ~4-6 months; a major-revision verdict adds more. Items marked (est.) have no published number.")
ws6.cell(row=2, column=1).font = note_font
ws6.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
ws6.row_dimensions[2].height = 60
headers6 = ["Venue", "Paper", "Time to first decision", "Accepted -> online", "Confidence / note"]
rows6 = [
    ["TMLR", "Identity", "Reviews <=4 weeks; decision ~2 months", "Days (OpenReview)",
     "Published policy. Fastest end-to-end; rolling submission."],
    ["Data-Centric Engineering (Cambridge)", "OT/ICS", "~6-8 weeks (est.)", "Fast; ~12 weeks submission->publication (median)",
     "12-week median is published; first-decision inferred. Unusually fast for an indexed journal."],
    ["Natural Language Processing (Cambridge)", "RegMap", "~3 months (target); one round reported 9.4 wks", "Weeks (Cambridge FirstView)",
     "Journal's stated target. Bimonthly."],
    ["Data & Policy (Cambridge)", "RegMap (alt)", "~2-3 months (est.)", "Weeks (FirstView)",
     "No official number published - estimate from typical CUP OA behaviour."],
    ["Network Science (Cambridge)", "Identity (alt)", "~2-3 months (est.)", "Weeks online; only 1 volume/year",
     "Estimate; weaker topical fit anyway."],
    ["ACM DTRAP", "Backup (any)", "~2-4 months (reviewers asked for 2 wks)", "Fast after acceptance",
     "Reviewer window published; total inferred. Rolling."],
    ["JSys (Journal of Systems Research)", "Backup", "~6 weeks to initial decision", "Fast (eScholarship)",
     "Journal's stated figure. Diamond OA."],
    ["USENIX CSET", "OT/ICS backup", "~6-8 weeks AFTER the annual deadline", "At the workshop (fixed date)",
     "You must wait for the next annual CFP deadline first - can be months before the clock starts. Confirm the current deadline."],
]
fills6 = [GREEN, GREEN, LIGHT, YELLOW, YELLOW, LIGHT, LIGHT, YELLOW]
end6 = write_table(ws6, 3, headers6, rows6, widths=[34, 18, 34, 34, 46], fills=fills6)
for i in range(len(rows6)):
    ws6.row_dimensions[4 + i].height = 50

# =========================================================================
out_dir = r"c:\Users\User\ai-cybersecurity-portfolio\docs\publishing"
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "publishing_recommendations.xlsx")
wb.save(out_path)
print("Saved:", out_path)
print("Sheets:", wb.sheetnames)
